from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from . import db
from .paths import EXPORT_DIR


def export_results_to_excel(election_id: int) -> Path:
    election = db.get_election(election_id)
    if not election:
        raise ValueError("Election not found")

    rows = db.get_results(election_id)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    safe_name = "".join(ch for ch in election["name"] if ch.isalnum() or ch in (" ", "-", "_")).strip()
    filename = f"{safe_name or 'election'}_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = EXPORT_DIR / filename

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Election", election["name"]])
    summary.append(["Exported At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary.append([])
    summary.append(["Poll", "Candidate", "Description", "Votes", "Winner"])

    by_poll: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        by_poll[row["poll_name"]].append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in summary[4]:
        cell.fill = header_fill
        cell.font = header_font

    for poll_name, poll_rows in by_poll.items():
        max_votes = max((r["votes"] for r in poll_rows), default=0)
        for row in poll_rows:
            winner = "Yes" if row["votes"] == max_votes and max_votes > 0 else ""
            summary.append([poll_name, row["candidate_name"], row["candidate_description"], row["votes"], winner])

        sheet = wb.create_sheet(title=poll_name[:31])
        sheet.append(["Candidate", "Description", "Votes", "Winner"])
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in poll_rows:
            winner = "Yes" if row["votes"] == max_votes and max_votes > 0 else ""
            sheet.append([row["candidate_name"], row["candidate_description"], row["votes"], winner])
        autosize(sheet)

    autosize(summary)
    wb.save(path)
    return path


def autosize(sheet) -> None:
    for column in sheet.columns:
        max_length = 0
        letter = column[0].column_letter
        for cell in column:
            max_length = max(max_length, len(str(cell.value or "")))
        sheet.column_dimensions[letter].width = min(max_length + 2, 48)
